import os
import sys
import re
from openpyxl import Workbook
from openpyxl import load_workbook
import gc

filename='ana.xlsx'


wb=load_workbook(filename=filename)
sheetnames=wb.get_sheet_names()
sheet=wb.get_sheet_by_name(sheetnames[0])

nwb=Workbook()
ws=nwb.active


for i in range(2,200,2):
    content = []
    for j in range(1,8):
        content.append(sheet.cell(row=i, column=j).value)
    content.append(sheet.cell(row=(i+1), column=5).value)
    content.append(sheet.cell(row=(i + 1), column=6).value)
    content.append(sheet.cell(row=(i + 1), column=7).value)
    ws.append(content)

nwb.save('ana2.xlsx')

