# -*- coding: utf-8 -*-
import read
from openpyxl import load_workbook
import numpy as np
from openpyxl import Workbook
import traceback
#站台规划
def handle(orderFilename,reFilename,orders, items,Me):
    #读取订单

    print('reading...')
    ordersData=read.readOrders(orderFilename,orders,items)
    reData=read.readResult(reFilename,Me)
    print('read finished')

    #统计item出现频次
    itemsAp=[[0,i] for i in range(items)]
    room=[0 for i in range(items)]
    for i in range(items):
        for j in range(orders):
            if ordersData[j][i]!=0:
                itemsAp[i][0]+=1
                room[i]+=ordersData[j][i]



    portNum=[[] for i in range(items)]
    #对出现频次进行降序
    newI=sorted(itemsAp, key=lambda itemsAp: itemsAp[0],reverse=True)
    rank=[[] for i in range(1500)]

    #进行架号编写前10%，10%-30%，余下
    ordersSq=[[] for i in range(orders)]
    portN=1
    jump=1
    s=0
    sumit=0
    left=400
    for i in range(items):
        # if(s==jump):
        #     s=0
        #     sumit=0
        #     portN+=1
        # if i>=int((items+1)*0.3):
        #     jump=4
        # elif i>=int((items+1)*0.1):
        #     jump=2
        sumit=room[newI[i][1]]
        if (sumit % 100 == 0):
            sumit = int(sumit / 100) * 100
        else:
            sumit = (int(sumit / 100) + 1) * 100
        if sumit>=left:
            while sumit>=left:
                portNum[newI[i][1]].append(portN)
                rank[portN].append((newI[i][1],left))
                portN+=1
                sumit=sumit-left
                left=400

            if sumit==0:
                continue
            else:
                left=left-sumit
                portNum[newI[i][1]].append(portN)
                rank[portN].append((newI[i][1],sumit))
        else:
            portNum[newI[i][1]].append(portN)
            rank[portN].append((newI[i][1], sumit))
            left = left - sumit
    # for i in range(items):
    #     print(newI[i][0],'  ',newI[i][1])
    #     print(portNum[newI[i][1]],room[newI[i][1]])
    print('arrange rank finished')

    for i in range(Me):
        for j in range(len(reData[i])):
            ordersSq[reData[i][j]].append(int(j/6)+1)
            ordersSq[reData[i][j]].append(i+1)

    wb1 = Workbook()
    ws1 = wb1.active
    title = ['rank','port','type','number']
    ws1.append(title)

    for i in range(1500):
        for x in rank[i]:
            content = []
            content.append(i)
            content.append(i)
            content.append(x[0])
            content.append(x[1])
            ws1.append(content)

    wb1.save('rank'+orderFilename)
    print('rank'+orderFilename,'created')
    wb = Workbook()
    ws = wb.active
    title=['订单次序','order','station','rank','item type','item quantity','rank left']
    ws.append(title)
    #i 是orede NUMBER,j 是items num
    leftsq=[[] for i in range(items)]
    pre=[]
    pre2=[]
    row=0
    maxsqe=0
    for i in range(1500):
        for x in rank[i]:
            leftsq[x[0]].append(x[1])
    for i in range(orders):

        for j in range(items):
            if ordersData[i][j]!=0:
                mm = ordersData[i][j]
                while True:
                    content = []
                    if maxsqe<=ordersSq[i][0]:
                        maxsqe=ordersSq[i][0]
                    content.append(ordersSq[i][0])
                    content.append('order' + str(i + 1))
                    content.append(ordersSq[i][1])
                    t = content
                    try:

                        if leftsq[j][0]>=mm:
                            if (len(pre2) > 2):
                                del pre2[0]
                            leftsq[j][0]=leftsq[j][0]-mm
                            t.append(portNum[j][0])
                            t.append(j)
                            t.append(mm)
                            t.append(leftsq[j][0])
                            if leftsq[j][0]==0:
                                del leftsq[j][0]
                                del portNum[j][0]
                            Fullcontent = t
                            ws.append(Fullcontent)
                            row+=1
                            break
                        else:

                            t.append(portNum[j][0])
                            t.append(j)
                            t.append(leftsq[j][0])
                            t.append(0)
                            mm=mm-leftsq[j][0]
                            Fullcontent = t
                            ws.append(Fullcontent)
                            row+=1
                            del leftsq[j][0]
                            del portNum[j][0]
                            if mm == 0:
                                break


                    except:
                        print('final'+orderFilename,'wrong!')
                        traceback.print_exc()
                        print('error')
                        return None
    wb.save('finalrandom'+orderFilename)
    return (row,maxsqe)
    print('finalrandom'+orderFilename,'created')









#handle('o3000tc100ac1.5an20.0.xlsx','reo3000tc100ac1.5an20.0.xlsx',3000,100)
