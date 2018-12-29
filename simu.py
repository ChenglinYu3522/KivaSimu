# -*- coding: utf-8 -*-
from openpyxl import load_workbook
from ARRANGE import arr
import random
from openpyxl import Workbook
global rankPosition
# rank position,creatre bt arrange.py
# station's position
global station

def getMission(step,rankPosition,takeresult):
    mission=[]


    for i in range(len(step)):
        for j in range(len(step[i])):
            if step[i][j]:
                mission.append((rankPosition[step[i][j][0]-1],step[i][j][0],j,takeresult[i][j][step[i][j][0]]))

                del step[i][j][0]
        if mission:
            break
    return mission

def simu(steps,sumL,filename,Me,maxX,maxY,vehicleNum):
    print('warehouse map arrange finished')
    rankPosition=arr(maxY,maxX)
    #[7, 1], [14, 1], [21, 1], [28, 1]
    station=[]
    stationInterval = int ((maxX - Me) / (Me + 1))
    positionX=0
    for i in range(Me):
        positionX+=stationInterval+1
        station.append([positionX,1])

    sumOn=0
    sumOff=0
    stationNum=Me

    f=open('f.txt','w')
    #sumL means the whole sqe times
    # sumL=5
    #steps means the whole steps
    # steps=202
    #filename means the step-file's name
    # filename='1.xlsx'
    #vehicle's num
    print('load simu data')
    workbook_ = load_workbook(filename=filename) #导入工作表
    sheetnames =workbook_.get_sheet_names() #获得表单名字
    sheet = workbook_.get_sheet_by_name(sheetnames[0]) #从工作表中提取某一表单
    #订单次序,order,station,rank,itemtype,item qu,rank left
    step=[[[] for j in range(stationNum)] for i in range(sumL)]
    take = [[[] for j in range (stationNum)] for i in range (sumL)]
    for rowNum in range(2,steps):
        step[sheet.cell(row=rowNum, column=1).value-1][sheet.cell(row=rowNum, column=3).value-1].append(sheet.cell(row=rowNum, column=4).value)
        take[sheet.cell(row=rowNum, column=1).value-1][sheet.cell(row=rowNum, column=3).value-1].append((sheet.cell(row=rowNum, column=4).value,sheet.cell(row=rowNum, column=6).value))

    takeresult=[[{} for j in range (stationNum) ] for i in range(sumL)]
    #step=[[[],[],[],[]],..[[],[],[],[]]] step[i][j] include a list,means in sqe i,the j station need which rank
    for i in range(sumL):
        for j in range(stationNum):
            for k in range(len(take[i][j])):
                if take[i][j][k][0] in takeresult[i][j]:
                    takeresult[i][j][take[i][j][k][0]]+=take[i][j][k][1]
                else:
                    takeresult[i][j][take[i][j][k][0]]=take[i][j][k][1]
            step[i][j]=list(set(step[i][j]))
    print('data load finished')

    print ('begin to simu...')
    #save the simu data as xlsx


    wb = Workbook()
    ws = wb.active

    ws.append(['vehicle编号','x','y','负载距离','空载距离','时间'])



    #rank State means rank's state is in use or free
    rankState=[0 for i in range(len(rankPosition))]
    rankWeight=[[500,500] for i in range(len(rankPosition))]
    #area's block

    area=[[False for j in range(maxY)]for i in range(maxX)]
    for i in range(maxX):
        if i%2==0:
            for j in range(7,maxY):
                area[i][j]=True

    #vehicle's state include position(now position and next position) and 3 states(move,at station,free),target position
    #and what rank it take now and waiting time,sumOn and sumOff distance
    #0 current position
    #1 next position
    #2 status
    #3 target position
    #4 carry rank rankID
    #5 stationWaitTime
    #6 sumOn_distance  sumOff_distance  sumTime
    vehicle=[[[1,i%4+2],[1,i%4+2],'free',[0,0],None,0,[0,0,0]] for i in range(vehicleNum)]

    #mission means the mission vehicle should do first
    mission=[]
    #t means the time
    t=0
    log=[]

    while True:

#newarea used to record the map of rank area
        newarea = [[False for j in range(maxY)] for i in range(maxX)]
        for i in range(maxX):
            if i % 2 == 0:
                for j in range(7, maxY):
                    newarea[i][j] = True
#use mission to get the mssions,let the vehicle to take mission
        if not mission:
#getMssion(step)  is a function to take some mission need to be done now
# ps.step means the what rank need to be take.
            mission=getMission(step,rankPosition,takeresult)
        for i in range(len(vehicle)):
            if vehicle[i][2]=='free' or (vehicle[i][2]=='move' and vehicle[i][4]==None and vehicle[i][3] in station):
                # if not mission:
                #     mission=getMission(step,rankPosition,takeresult)

                if mission:
                    for j in range(len(mission)):

                        if rankState[mission[j][1]-1]!=0:
                            continue
                        vehicle[i][2]='move'
                        vehicle[i][3]=mission[j][0]
                        rankState[mission[j][1]-1]=mission[j][2]+1
                        rankWeight[mission[j][1]-1][1]=rankWeight[mission[j][1]-1][0]-mission[j][3]
                        del mission[j]

                        break
                    # if vehicle[i][2]=='free':
                    #     mission=mission+getMission(step,rankPosition)
                    #
                    #     for j in range(len(mission)):
                    #
                    #         if rankState[mission[j][1]-1]!=0:
                    #             continue
                    #         vehicle[i][2]='move'
                    #         vehicle[i][3]=mission[j][0]
                    #         rankState[mission[j][1]-1]=mission[j][2]+1
                    #         del mission[j]
                    #
                    #         break
                    # missionLen = len (mission)
                    # newmissionLen=-1
                    # while (vehicle[i][2]=='free'):
                    #     if (missionLen==newmissionLen):
                    #         break
                    #     missionLen = len (mission)
                    #     for j in range(len(mission)):
                    #         if rankState[mission[j][1]-1]!=0:
                    #             continue
                    #         vehicle[i][2]='move'
                    #         vehicle[i][3]=mission[j][0]
                    #         rankState[mission[j][1]-1]=mission[j][2]+1
                    #         del mission[j]
                    #         break
                    #     if vehicle[i][2]=='move':
                    #         break
                    #     else:
                    #         mission = mission + getMission (step,rankPosition)
                    #         newmissionLen=len(mission)

#if not all vechile free means the mission is going to be finished,now let the left vehicle go out.
        for i in range(len(vehicle)):
            if vehicle[i][2]=='free' and vehicle[i][0] not in station:
                vehicle[i][2]='move'
                vehicle[i][3]=station[i%len(station)]




#predict every vehicle's next posiyion
#use A* al,at first think about to close the target,then
        for v in vehicle:
            newarea[v[0][0]][v[0][1]]=True



        for i in range(len(vehicle)):
            for j in range(i):
                newarea[vehicle[j][1][0]][vehicle[j][1][1]] = True
            for w in station:
                newarea[w[0]][w[1]]=False
            if vehicle[i][2]=='free':
                vehicle[i][6][2]+=1
                continue
            if vehicle[i][2]=='station':
                vehicle[i][5]+=1
                if vehicle[i][5]==5:
                    rankWeight[vehicle[i][4]-1][0]=rankWeight[vehicle[i][4]-1][1]
                    vehicle[i][2] = 'move'
                    vehicle[i][5]= 0
                    vehicle[i][3]= rankPosition[vehicle[i][4]-1]
                continue
            if vehicle[i][2]=='move':
                if vehicle[i][0]==vehicle[i][3]:
                    if vehicle[i][4]==None and vehicle[i][3] in rankPosition:
                        vehicle[i][4]=rankPosition.index(vehicle[i][3])+1

                        vehicle[i][3]=station[rankState[rankPosition.index(vehicle[i][3])]-1]

                        continue
                    elif vehicle[i][4]!=None and vehicle[i][3] in rankPosition:
                        argetPo=[0,6]
                        if vehicle[i][3][0]%4==2:
                            argetPo[0]=vehicle[i][3][0]+1
                        else:
                            argetPo[0] = vehicle[i][3][0]-1
                        rankState[vehicle[i][4] - 1] = 0
                        vehicle[i][4]=None
                        vehicle[i][3]=argetPo

                        continue
                    elif vehicle[i][4]!=None and vehicle[i][3] not in rankPosition:
                        vehicle[i][2] = 'station'
                        continue
                    else:
                        vehicle[i][2] = 'free'
                        continue
                else:
                    now=vehicle[i][0]
                    tar=vehicle[i][3]
                    dir=None


                    if tar[1]>6:
                        if now[1]<=6 and tar[0]%4==2:
                            if now[0]<(tar[0]-1):
                                dir='right'
                            elif now[0]==(tar[0]-1):
                                dir='up'
                            else:
                                dir='left'
                        elif now[1]<=6 and tar[0]%4==0:
                            if now[0]<(tar[0]+1):
                                dir='right'
                            elif now[0]==(tar[0]+1):
                                dir='up'
                            else:
                                dir='left'
                        else:
                            if now[1]==tar[1]:
                                if tar[0]%4==2:
                                    dir='right'
                                else:
                                    dir='left'
                            else:
                                dir='up'
                    elif tar[1]<=6:
                        if now[1]>6:
                            if now[0]%4==2:
                                dir='right'
                            elif now[0]%4==0:
                                dir='left'
                            else:
                                dir='down'
                        if now[1]<=6:
                            canDownx=tar[0]
                            while(True):
                                if canDownx%4==3 or canDownx%4==0:
                                    break
                                canDownx-=1


                            if now[0]==canDownx and now[1]>tar[1]:
                                dir='down'
                                # if now[0]%4==3 or now[0]%4==0:
                                #     dir='down'
                                # else:
                                #     if now[1]%2==1:
                                #         dir='right'
                                #     else:
                                #         dir='left'
                            elif now[0]<canDownx and now[1]>tar[1]:
                                # if now[0] % 4 == 3 or now[0] % 4 == 0:
                                #     dir='down'
                                # else:
                                #     dir='right'
                                if now[0] % 4==3 or now[0]%4==0:
                                    dir='down'
                                else:
                                    dir='right'
                                # if tar[1]%2==1:
                                #     if now[0]%4==3 or now[0]%4==0:
                                #         dir='down'
                                #     else:
                                #         dir='right'
                                # else:
                                #     dir='right'
                            elif now[0]>canDownx and now[1]>tar[1]:
                                # if now[0] % 4 == 3 or now[0] % 4 == 0:
                                #     dir='down'
                                # else:
                                #     dir='left'
                                dir='left'
                                # if tar[1] % 2 == 1:
                                #     dir='left'
                                # else:
                                #     if now[0]%4==1 or now[0]%4==2:
                                #         dir='down'
                                #     else:
                                #         dir='right'
                            elif now[0]<tar[0] and now[1]==tar[1]:
                                dir='right'
                            # elif now[0]>tar[0] and now[1]==tar[1]:
                            #     dir='left'
                    if now==tar:
                        dir=None





                next = [vehicle[i][0][0], vehicle[i][0][1]]
                if dir=='down':

                    if newarea[vehicle[i][0][0]][vehicle[i][0][1]-1]==True:
                        next=[vehicle[i][0][0],vehicle[i][0][1]]
                    else:
                        next = [vehicle[i][0][0], vehicle[i][0][1] - 1]
                        if vehicle[i][4]:
                            vehicle[i][6][0]+=1
                        else:
                            vehicle[i][6][1]+=1
                elif dir=='up':

                    if newarea[vehicle[i][0][0]][vehicle[i][0][1]+1]==True:
                        next=[vehicle[i][0][0],vehicle[i][0][1]]
                    else:
                        next = [vehicle[i][0][0], vehicle[i][0][1] + 1]
                        if vehicle[i][4]:
                            vehicle[i][6][0]+=1
                        else:
                            vehicle[i][6][1]+=1
                elif dir=='right':

                    if vehicle[i][0][0]+1==vehicle[i][3][0] and vehicle[i][0][1]==vehicle[i][3][1]:
                        next = [vehicle[i][0][0] + 1, vehicle[i][0][1]]
                        if vehicle[i][4]:
                            vehicle[i][6][0] += 1
                        else:
                            vehicle[i][6][1] += 1
                    else:
                        if vehicle[i][0][1]>6:
                            if newarea[vehicle[i][0][0]+1][vehicle[i][0][1] ] == True:
                                next = [vehicle[i][0][0], vehicle[i][0][1]]
                            else:
                                next = [vehicle[i][0][0]+1, vehicle[i][0][1]]
                                if vehicle[i][4]:
                                    vehicle[i][6][0] += 1
                                else:
                                    vehicle[i][6][1] += 1
                        else:
                            if newarea[vehicle[i][0][0] + 1][vehicle[i][0][1]] == True:
                                if vehicle[i][0][1]%2==1:
                                    next = [vehicle[i][0][0], vehicle[i][0][1]]
                                else:
                                    if (vehicle[i][0][0]%4==1 or vehicle[i][0][0]%4==2) and vehicle[i][0][1]<6:
                                        if newarea[vehicle[i][0][0]][vehicle[i][0][1]+1] == True:
                                            next = [vehicle[i][0][0], vehicle[i][0][1]]
                                        else:
                                            next = [vehicle[i][0][0], vehicle[i][0][1] + 1]
                                            if vehicle[i][4]:
                                                vehicle[i][6][0] += 1
                                            else:
                                                vehicle[i][6][1] += 1
                                    elif (vehicle[i][0][0]%4==3 or vehicle[i][0][0]%4==0):
                                        if newarea[vehicle[i][0][0]][vehicle[i][0][1]-1] == True:
                                            next = [vehicle[i][0][0], vehicle[i][0][1]]
                                        else:
                                            next = [vehicle[i][0][0], vehicle[i][0][1] - 1]
                                            if vehicle[i][4]:
                                                vehicle[i][6][0] += 1
                                            else:
                                                vehicle[i][6][1] += 1
                                    elif (vehicle[i][0][0]%4==1 or vehicle[i][0][0]%4==2) and vehicle[i][0][1]==6:
                                        if newarea[vehicle[i][0][0]-1][vehicle[i][0][1]] == True:
                                            next = [vehicle[i][0][0], vehicle[i][0][1]]
                                        else:
                                            next = [vehicle[i][0][0]-1, vehicle[i][0][1]]
                                            if vehicle[i][4]:
                                                vehicle[i][6][0] += 1
                                            else:
                                                vehicle[i][6][1] += 1

                                    else:
                                        if newarea[vehicle[i][0][0] - 1][vehicle[i][0][1]] == True:
                                            next = [vehicle[i][0][0], vehicle[i][0][1]]
                                        else:
                                            next = [vehicle[i][0][0] - 1, vehicle[i][0][1]]
                                            if vehicle[i][4]:
                                                vehicle[i][6][0] += 1
                                            else:
                                                vehicle[i][6][1] += 1
                            else:
                                if vehicle[i][0][1]%2==1:
                                    next = [vehicle[i][0][0]+1, vehicle[i][0][1]]
                                    if vehicle[i][4]:
                                        vehicle[i][6][0] += 1
                                    else:
                                        vehicle[i][6][1] += 1
                                else:
                                    if (vehicle[i][0][0]%4==1 or vehicle[i][0][0]%4==2) and vehicle[i][0][1]<6:
                                        if newarea[vehicle[i][0][0]][vehicle[i][0][1]+1] == True:
                                            next = [vehicle[i][0][0], vehicle[i][0][1]]
                                        else:
                                            next = [vehicle[i][0][0], vehicle[i][0][1] + 1]
                                            if vehicle[i][4]:
                                                vehicle[i][6][0] += 1
                                            else:
                                                vehicle[i][6][1] += 1
                                    elif (vehicle[i][0][0]%4==1 or vehicle[i][0][0]%4==2) and vehicle[i][0][1]==6:
                                        if newarea[vehicle[i][0][0]-1][vehicle[i][0][1]] == True:
                                            next = [vehicle[i][0][0], vehicle[i][0][1]]
                                        else:
                                            next = [vehicle[i][0][0]-1, vehicle[i][0][1]]
                                            if vehicle[i][4]:
                                                vehicle[i][6][0] += 1
                                            else:
                                                vehicle[i][6][1] += 1
                                    elif (vehicle[i][0][0]%4==3 or vehicle[i][0][0]%4==0):
                                        if newarea[vehicle[i][0][0]][vehicle[i][0][1]-1] == True:
                                            next = [vehicle[i][0][0], vehicle[i][0][1]]
                                        else:
                                            next = [vehicle[i][0][0], vehicle[i][0][1] - 1]
                                            if vehicle[i][4]:
                                                vehicle[i][6][0] += 1
                                            else:
                                                vehicle[i][6][1] += 1
                                    else:
                                        if newarea[vehicle[i][0][0] - 1][vehicle[i][0][1]] == True:
                                            next = [vehicle[i][0][0], vehicle[i][0][1]]
                                        else:
                                            next = [vehicle[i][0][0] - 1, vehicle[i][0][1]]
                                            if vehicle[i][4]:
                                                vehicle[i][6][0] += 1
                                            else:
                                                vehicle[i][6][1] += 1



                elif dir=='left':
                    if vehicle[i][0][0]-1==vehicle[i][3][0] and vehicle[i][0][1]==vehicle[i][3][1]:
                        next = [vehicle[i][0][0] - 1, vehicle[i][0][1]]
                        if vehicle[i][4]:
                            vehicle[i][6][0] += 1
                        else:
                            vehicle[i][6][1] += 1
                    else:
                        if vehicle[i][0][1]>6:
                            if newarea[vehicle[i][0][0]-1][vehicle[i][0][1] ] == True:
                                next = [vehicle[i][0][0], vehicle[i][0][1]]
                            else:
                                next = [vehicle[i][0][0]-1, vehicle[i][0][1]]
                                if vehicle[i][4]:
                                    vehicle[i][6][0] += 1
                                else:
                                    vehicle[i][6][1] += 1
                        else:
                            if newarea[vehicle[i][0][0] - 1][vehicle[i][0][1]] == True:
                                if vehicle[i][0][1]%2==0:
                                    next = [vehicle[i][0][0], vehicle[i][0][1]]
                                else:
                                    if (vehicle[i][0][0]%4==1 or vehicle[i][0][0]%4==2):
                                        if newarea[vehicle[i][0][0]][vehicle[i][0][1]+1] == True:
                                            next = [vehicle[i][0][0], vehicle[i][0][1]]
                                        else:
                                            next = [vehicle[i][0][0], vehicle[i][0][1]+1]
                                            if vehicle[i][4]:
                                                vehicle[i][6][0] += 1
                                            else:
                                                vehicle[i][6][1] += 1
                                    elif (vehicle[i][0][0]%4==3 or vehicle[i][0][0]%4==0) and vehicle[i][0][1]>1:
                                        if newarea[vehicle[i][0][0]][vehicle[i][0][1]-1] == True:
                                            next = [vehicle[i][0][0], vehicle[i][0][1]]
                                        else:
                                            next = [vehicle[i][0][0], vehicle[i][0][1] - 1]
                                            if vehicle[i][4]:
                                                vehicle[i][6][0] += 1
                                            else:
                                                vehicle[i][6][1] += 1
                                    elif (vehicle[i][0][0]%4==3 or vehicle[i][0][0]%4==0) and vehicle[i][0][1]==1:
                                        if newarea[vehicle[i][0][0]+1][vehicle[i][0][1]] == True:
                                            next = [vehicle[i][0][0], vehicle[i][0][1]]
                                        else:
                                            next = [vehicle[i][0][0]+1, vehicle[i][0][1]]
                                            if vehicle[i][4]:
                                                vehicle[i][6][0] += 1
                                            else:
                                                vehicle[i][6][1] += 1
                                    else:
                                        if newarea[vehicle[i][0][0] + 1][vehicle[i][0][1]] == True:
                                            next = [vehicle[i][0][0], vehicle[i][0][1]]
                                        else:
                                            next = [vehicle[i][0][0] + 1, vehicle[i][0][1]]
                                            if vehicle[i][4]:
                                                vehicle[i][6][0] += 1
                                            else:
                                                vehicle[i][6][1] += 1
                            else:
                                if vehicle[i][0][1]%2==0:
                                    next = [vehicle[i][0][0]-1, vehicle[i][0][1]]
                                    if vehicle[i][4]:
                                        vehicle[i][6][0] += 1
                                    else:
                                        vehicle[i][6][1] += 1
                                else:
                                    if (vehicle[i][0][0]%4==1 or vehicle[i][0][0]%4==2) :
                                        if newarea[vehicle[i][0][0]][vehicle[i][0][1]+1] == True:
                                            next = [vehicle[i][0][0], vehicle[i][0][1]]
                                        else:
                                            next = [vehicle[i][0][0], vehicle[i][0][1] + 1]
                                            if vehicle[i][4]:
                                                vehicle[i][6][0] += 1
                                            else:
                                                vehicle[i][6][1] += 1
                                    elif (vehicle[i][0][0]%4==3 or vehicle[i][0][0]%4==0) and vehicle[i][0][1]>1:
                                        if newarea[vehicle[i][0][0]][vehicle[i][0][1]-1] == True:
                                            next = [vehicle[i][0][0], vehicle[i][0][1]]
                                        else:
                                            next = [vehicle[i][0][0], vehicle[i][0][1] - 1]
                                            if vehicle[i][4]:
                                                vehicle[i][6][0] += 1
                                            else:
                                                vehicle[i][6][1] += 1
                                    elif (vehicle[i][0][0]%4==3 or vehicle[i][0][0]%4==0) and vehicle[i][0][1]==1:
                                        if newarea[vehicle[i][0][0]+1][vehicle[i][0][1]] == True:
                                            next = [vehicle[i][0][0], vehicle[i][0][1]]
                                        else:
                                            next = [vehicle[i][0][0]+1, vehicle[i][0][1]]
                                            if vehicle[i][4]:
                                                vehicle[i][6][0] += 1
                                            else:
                                                vehicle[i][6][1] += 1
                                    else:
                                        if newarea[vehicle[i][0][0] + 1][vehicle[i][0][1]] == True:
                                            next = [vehicle[i][0][0], vehicle[i][0][1]]
                                        else:
                                            next = [vehicle[i][0][0] + 1, vehicle[i][0][1]]
                                            if vehicle[i][4]:
                                                vehicle[i][6][0] += 1
                                            else:
                                                vehicle[i][6][1] += 1

                vehicle[i][1]=next






        for i in range(len(vehicle)):

            content=[i+1,vehicle[i][0][0],vehicle[i][0][1],vehicle[i][6][0],vehicle[i][6][1],t]
            if vehicle[i][4]:
                content.append(rankWeight[vehicle[i][4]-1][0])
            else:
                content.append(0)
            ws.append(content)
            sumOn+=vehicle[i][6][0]
            sumOff+=vehicle[i][6][1]
            # print((vehicle[i][0],vehicle[i][1],vehicle[i][3]),end='')
            vehicle[i][0]=vehicle[i][1]
        # print()


        t += 1


        if not mission:
            out=True
            for v in vehicle:
                if v[2]!='free':
                    out=False
            mission = mission + getMission(step,rankPosition,takeresult)
            if mission:
                out = False
            if out:
                break

        nowarea = [[False for j in range(47)] for i in range(36)]
        for i in range(36):
            if i % 2 == 0:
                for j in range(7, 47):
                    nowarea[i][j] = True
        for x in vehicle:
            nowarea[x[0][0]][x[0][1]]=True

    wb.save('simu'+filename)
    print('simu'+filename+' fininshed')
    print('---------------------------------one group end--------------------------------------')
    print()
    print()
    print()











































