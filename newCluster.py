# -*- coding: utf-8 -*-
#聚类

from openpyxl import load_workbook
import numpy as np
from openpyxl import Workbook

# filename=str(input('input the head filename'))
# sheets=str(input('input the num of sheets'))


def createSheets(filename,orders,items,orderseach,Me):
    print('begin to arrange station..')

    #orderseach=12;
    plat=[False for i in range(Me)]
    platstat=[[] for i in range(Me)]
    c=0;
    platnum=[orderseach for i in range(Me)]
    workbook_ = load_workbook(filename=filename) #导入工作表
    sheetnames =workbook_.get_sheet_names() #获得表单名字
    sheet = workbook_.get_sheet_by_name(sheetnames[0]) #从工作表中提取某一表单

    sumdata=[]
    sss=[[] for i in range(Me)]
    f=open('same' + filename.strip('xlsx') + 'txt', 'r')
    same=eval(f.read())
    #orders+2
    for colNum in range(2,orders+2):
        data=[]
        #items+2
        for rowNum in range(2,items+2):
            i=0
            if(sheet.cell(row=rowNum, column=colNum).value):
                i=1
            data.append(i) #获得数据

        for j in range(len(same)):
            if data[same[j][0]]==1:
                data[same[j][1]]=1
        sumdata.append(data)



    # print('first state:\n',np.mat(sumdata))

    left=orders
    inOrder=[False for i in range(orders)]


    while(True):
        if(c>(Me-1)):
            # print('clean the plat')
            plat = [False for i in range(Me)]
            platstat = [[] for i in range(Me)]
            platnum = [orderseach for i in range(Me)]
            c=0
        if(not plat[c]):
            if(left==0):
                break
            if(left<orderseach):
                for i in range(orders):
                    if(not inOrder[i]):
                        platstat[c].append(i)
                        sss[c].append(i)
                        inOrder[i]=True
                        left=0
                # for i in range(Me):
                #     # print('the',i,'th plat ordernum:',platstat[i])
                break
            else:
                X = np.mat(sumdata)

                S = X * X.T

                a = S.getA()
                raw, column = S.shape
                for i in range(raw):
                    a[i][i] = -1
                S = np.mat(a)
                # print('the sim:\n', S)
                _positon = np.argmax(S)
                m, n = divmod (_positon, column)
                while (inOrder[m]):
                    m=(1+m)%orders
                while (inOrder[n]):
                    n=(n+1)%orders

                inOrder[m] = True
                platstat[c].append(m)
                platstat[c].append(n)
                sss[c].append(m)
                sss[c].append(n)
                inOrder[n]=True
                left=left-2

                platnum[c]-=2
                # for i in range(Me):
                #     print('the',i,'th plat ordernum:',platstat[i])
                k = []
                for i in range(len(sumdata[0])):
                    if (sumdata[m][i] or sumdata[n][i]):
                        sumdata[m][i] = 0
                        sumdata[n][i] = 0
                        k.append(1)
                    else:
                        k.append(0)
                P = np.mat(k)
                X = np.mat(sumdata)
                while(platnum[c]):
                    if(left==0):
                        break;

                    S = X * P.T
                    _positon = np.argmax(S)
                    # print(S.T)
                    m, n = divmod(_positon, column)
                    # print(n)
                    while (inOrder[n]):
                        n = (n+1)%orders

                    platstat[c].append(n)
                    sss[c].append(n)
                    inOrder[n]=True

                    for i in range(len(sumdata[n])):
                        sumdata[n][i]=0
                    X = np.mat(sumdata)
                    left-=1
                    platnum[c]-=1
                # for i in range(Me):
                #     print('the',i,'th plat ordernum:',platstat[i])
                plat[c]=True
                c+=1;

    wb = Workbook()
    ws = wb.active

    for i in range(Me):

        ws.append(sss[i])

    wb.save('Newre'+filename)
    print('Newre'+filename,'created')