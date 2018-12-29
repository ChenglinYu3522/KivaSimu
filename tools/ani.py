from matplotlib import pyplot as plt
from matplotlib import animation
import matplotlib.patches as patches
from openpyxl import load_workbook


def ana(filenamex,X,Y,Me):
    workbook_ = load_workbook(filename=filenamex) #导入工作表
    sheetnames =workbook_.get_sheet_names() #获得表单名字
    sheet = workbook_.get_sheet_by_name(sheetnames[0]) #从工作表中提取某一表单

    vechile=[[] for i in range(938)]
    row=2
    z=-1
    t=0
    while row<295:
        if t%6==0:
            z+=1
            t=0
        vechile[z].append([sheet.cell(row=row, column=2).value-1,sheet.cell(row=row, column=3).value-1,sheet.cell(row=row, column=4).value,sheet.cell(row=row, column=5).value])
        t+=1
        row+=1


    print(vechile)











    #我们的数据是一个0~2π内的正弦曲线
    fig,ax1 = plt.subplots()
    plt.xlim((0, X))
    plt.ylim((0, Y))

    global color,re
    color='green'
    re='green'


    for i in range(X):
        if i%2==1:
            ax1.add_patch(
                patches.Rectangle(
                    (i, 6),   # (x,y)
                    1,          # width
                    40,          # height
                )
            )

    for j in vechile[0]:
        ax1.add_patch(
            patches.Rectangle(
                    (j[0], j[1]),  # (x,y)
                    1,  # width
                    1,  # height
                    facecolor=color
                )

            )
    station=[]
    stationInterval = int ((X - Me) / (Me + 1))
    positionX = 0
    for i in range (Me):
        positionX += stationInterval + 1
        station.append ([positionX, 1])
    for ci in station:
        ax1.add_patch(
            patches.Rectangle(
                (ci[0]-1, ci[1]-1),  # (x,y)
                1,  # width
                1,  # height
                facecolor='yellow'
            )

        )

        # ax1.add_patch(
        #     patches.Rectangle(
        #         (13, 0),  # (x,y)
        #         1,  # width
        #         1,  # height
        #         facecolor='yellow'
        #     )
        #
        # )
        #
        #
        # ax1.add_patch(
        #     patches.Rectangle(
        #         (20, 0),  # (x,y)
        #         1,  # width
        #         1,  # height
        #         facecolor='yellow'
        #     )
        #
        # )
        #
        # ax1.add_patch(
        #     patches.Rectangle(
        #         (27, 0),  # (x,y)
        #         1,  # width
        #         1,  # height
        #         facecolor='yellow'
        #     )
        #
        # )





    #接着，构造自定义动画函数animate，用来更新每一帧上各个x对应的y坐标值，参数表示第i帧
    def animate(i):
        global color, re
        plt.cla()
        plt.xlim((0, X))
        plt.ylim((0, Y))


        for i1 in range(X):
            if i1 % 2 == 1:
                ax1.add_patch(
                    patches.Rectangle(
                        (i1, 6),  # (x,y)
                        1,  # width
                        40,  # height
                        facecolor='blue'

                    )
                )


        if i%2==0:
            if i==0:
                return ax1,

            z=0
            for j in vechile[int(i/2)]:
                color=re
                if vechile[int(i/2)-1][z][3]<j[3]:
                    color='green'
                    re='green'
                if vechile[int(i/2)-1][z][2]<j[2]:
                    color='red'
                    re='red'
                z+=1

                ax1.add_patch(
                    patches.Rectangle(
                        (j[0], j[1]),  # (x,y)
                        1,  # width
                        1,  # height
                        facecolor=color
                    )

                )
        else:
            z=0

            for j in vechile[int(i / 2)]:

                color = re
                # if vechile[int(i / 2) - 1][z][3] < j[3]:
                #     color = 'green'
                #     re = 'green'
                # if vechile[int(i / 2) - 1][z][2] < j[2]:
                #     color = 'red'
                #     re = 'red'
                if vechile[int(i/2+1)][z][0]>j[0]:
                    x=j[0]+0.5
                elif vechile[int(i/2+1)][z][0]<j[0]:
                    x=j[0]-0.5
                else:
                    x=j[0]
                if vechile[int(i/2+1)][z][1]>j[1]:
                    y=j[1]+0.5
                elif vechile[int(i/2+1)][z][1]<j[1]:
                    y=j[1]-0.5
                else:
                    y=j[1]
                z+=1
                ax1.add_patch(
                    patches.Rectangle(
                        (x,y),  # (x,y)
                        1,  # width
                        1,  # height
                        facecolor=color
                    ))

        ax1.add_patch(
            patches.Rectangle(
                (6, 0),  # (x,y)
                1,  # width
                1,  # height
                facecolor='yellow'
            )

        )

        ax1.add_patch(
            patches.Rectangle(
                (13, 0),  # (x,y)
                1,  # width
                1,  # height
                facecolor='yellow'
            )

        )

        ax1.add_patch(
            patches.Rectangle(
                (20, 0),  # (x,y)
                1,  # width
                1,  # height
                facecolor='yellow'
            )

        )

        ax1.add_patch(
            patches.Rectangle(
                (27, 0),  # (x,y)
                1,  # width
                1,  # height
                facecolor='yellow'
            )

        )







        return ax1,
    #然后，构造开始帧函数init
    def init():
        for j in vechile[0]:
            ax1.add_patch(
                patches.Rectangle(
                    (j[0], j[1]),  # (x,y)
                    1,  # width
                    1,  # height
                    facecolor='coral'
                )

            )
        for i in range(36):
            if i % 2 == 1:
                ax1.add_patch(
                    patches.Rectangle(
                        (i, 6),  # (x,y)
                        1,  # width
                        40,  # height

                    )

                )

        return ax1,


    #fig 进行动画绘制的figure
    #func 自定义动画函数，即传入刚定义的函数animate
    #frames 动画长度，一次循环包含的帧数
    #init_func 自定义开始帧，即传入刚定义的函数init
    #interval 更新频率，以ms计
    #blit 选择更新所有点，还是仅更新产生变化的点。应选择True，但mac用户请选择False，否则无法显示动画

    ani = animation.FuncAnimation(fig=fig,
                                  func=animate,
                                  frames=1500,
                                  init_func=init,
                                  interval=20,
                                  blit=False)
    # plt.show()


    ani.save('b.mp4', fps=5)