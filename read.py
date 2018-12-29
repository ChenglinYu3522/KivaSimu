# -*- coding: utf-8 -*-
from openpyxl import load_workbook
import numpy as np
from openpyxl import Workbook
import copy

def readOrders(filename,orders,items):
    workbook_ = load_workbook(filename=filename) #导入工作表
    sheetnames =workbook_.get_sheet_names() #获得表单名字
    sheet = workbook_.get_sheet_by_name(sheetnames[0]) #从工作表中提取某一表单
    sumdata=[]
    for colNum in range(2,orders+2):
        data=[]
        for rowNum in range(2,items+2):
            data.append(sheet.cell(row=rowNum, column=colNum).value) #获得数据
        sumdata.append(data)
    return sumdata


def readResult(filename,Me):
    workbook_ = load_workbook(filename=filename) #导入工作表
    sheetnames =workbook_.get_sheet_names() #获得表单名字
    sheet = workbook_.get_sheet_by_name(sheetnames[0]) #从工作表中提取某一表单
    data=[[] for i in range(Me)]
    for rowNum in range(1,Me+1):
        colNum=1
        while(sheet.cell(row=rowNum, column=colNum).value!=None):
            data[rowNum-1].append(sheet.cell(row=rowNum, column=colNum).value)
            colNum+=1
    return data

