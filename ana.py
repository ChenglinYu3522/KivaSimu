import os
import sys
import re
from openpyxl import Workbook
from openpyxl import load_workbook
import gc
location=None
filename=[]
for loa,b,files in os.walk(sys.path[0]):
    location=loa
    if files:
        filename=files
        break

aa=[]
bb=[]
for file in filename:
    if 'new' in file:
        aa.append(file)



for file in filename:
    for a in aa:
        if file==''.join(map(str,a.split('new'))) and file!=a:
            bb.append(file)

filename=aa+bb



openwb=Workbook()
ws=openwb.active


num=0
for names in filename:
    if 'simu' not in names:
        continue
    else:
        a=0
        if 'new' in names:
            a=1
        content=re.split('[a-z|A-z|.]+',names)
        wb = load_workbook(filename=names)
        sheetnames = wb.get_sheet_names()
        sheet = wb.get_sheet_by_name(sheetnames[0])
        row = sheet.max_row - 2
        load=sheet.cell(row=row, column=4).value+sheet.cell(row=row+1, column=4).value+sheet.cell(row=row+2, column=4).value
        free = sheet.cell(row=row, column=5).value + sheet.cell(row=row + 1, column=5).value + sheet.cell(row=row + 2,column=5).value
        del wb, sheet
        gc.collect()
        content.append(load)
        content.append(free)
        content.append(a)
        ws.append(content)
        num+=1
        print(num)

openwb.save('ana.xlsx')